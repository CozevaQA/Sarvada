import base64
import os
import shutil

import pyautogui
import re
import time
from datetime import datetime, timedelta
from dateutil import parser
from openpyxl.styles import PatternFill,Font
### Write in such a way that you can always edit the page you will compare for time load
from selenium import webdriver
from selenium.common import ElementNotInteractableException, ElementClickInterceptedException, NoSuchElementException, \
    TimeoutException, StaleElementReferenceException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from Utilities import get_credentials
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

# initialize Webdriver
#service=ChromeService(ChromeDriverManager().install())
download_dir="C:\\Users\\ssrivastava\\Downloads"
report="C:\\VerificationReports\\"
def setup():
    try:
        chrome_profile_path = "C:\\chromeprofile"
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-notifications")
        options.add_argument("--start-maximized")
        options.add_experimental_option("detach", True)
        options.add_experimental_option("prefs", {
            "download.default_directory": download_dir,
            "download.prompt_for_download":False,
            "safebrowsing.enabled": True
        })

        options.add_argument(r"--user-data-dir=C:\\chromeprofile")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
        print("Chrome driver set up successful ")
        driver.get("https://www.google.com")
    except Exception as e:
        print(type(e).__name__)
    return driver

def secure_messaging_page_getURL(customer_id):
    customer_list_url = []
    sm_customer_id = customer_id
    session_var = 'app_id=cozeva_messages&custId=' + str(sm_customer_id) + '&payerId=' + str(
        sm_customer_id) + '&orgId=' + str(sm_customer_id)+'&vgpId=' + str(sm_customer_id)+'&vpId=' + str(sm_customer_id)
    encoded_string = base64.b64encode(session_var.encode('utf-8'))
    customer_list_url.append(encoded_string)
    for idx, val in enumerate(customer_list_url):
        url=("https://cert.cozeva.com/cozeva_messages/?session=" + val.decode('utf-8'))
    return url

def wait_to_load(driver,timeout):
    loader_element_class='ajax_preloader'
    WebDriverWait(driver,timeout).until(EC.invisibility_of_element_located((By.CLASS_NAME, loader_element_class)))




def is_valid_datetime(dt_str):
    try:
        parser.parse(dt_str)
        return True
    except ValueError:
        return False

#openpyxl function to save worksheet by applying formatting to headers and pass fail column
#def worksheet_format_and_save(workbook,worksheet_name):


#Checks contact log from secure messaging URL and writes status to report - worksheet , role
def check_contact_log(driver,workbook,sample_logger,run_from,customer_ids):
    # xpaths

    dropdown_icon_xpath = '//li[@id="messaging_stickets_access"]//child::i[@class="material-icons header-toggle"]'
    contact_log_link_xpath = '//*[text()="Contact Log"]'
    first_column_xpath = '//th[text()="Patient"]'
    records_xpath = '//tbody//tr'
    no_of_records_xpath='//div[@class="dataTables_info"]'
    columns_xpath='//thead//tr//th'
    list_of_patient_xpath='//tbody//tr//child::a'
    next_xpath='//button[text()="Next"]'
    previous_xpath = '//button[text()="Previous"]'
    name_column_xpath='//tbody//tr//child::a'
    filter_xpath='//i[text()="filter_list"]'
    patient_name_placeholder_xpath='//input[@title="Patient"]'
    apply_xpath='//a[@class="rfloat green-text datatable_apply" and text()="Apply"]'
    file_download = '//a[text()="file_download"]'
    visible_UL = '//ul[@id="datatable_bulk_filter_0_contact_log"]'
    link_to_click = '//a[text()="Export all to CSV "]'
    stickets_alert = '//div[@id="stickets_alert"]'
    yes_accept = '//div[@id = "stickets_alert"]//child::a[text() = "YES"]'
    toast_message_appears = '//div[ @ class = "toast green green-text lighten-5 text-darken-3 drupal_message drupal_status_message"]'
    encounter_dos_xpath='//tbody//child::tr[@role="row"]//child::td[5]'
    dos_dos_xpath='//tbody//child::tr[@role="row"]//child::td[14]'
    no_actual_record_xpath='//*[text()="No matching records found"]'
    notification_icon = '//a[@id="notification_button"]//span'
    no_of_notifications = '//a[@id ="notification_button"]//span'
    notification_panel = '//div[@id="notification_panel" and @class="dropdown-content loaded"]'
    notification_received= '//div[@class ="notif-message " and text()="Your request for all Contact Log Extract has been completed."]'
    link_to_download_click='//a[@style="color: #114C7F; font-weight: bold;"]'
    banner_remove_xpath='//div[@class="banner_footer"]//child::a[text()="Hide"]'



    for customer_id in customer_ids:
        # create worksheet and add headers
        sheet = workbook.create_sheet('Contact Log')
        headers = ["Test Area", "Test Case", "Status", "Comments/Time taken(in s)"]
        sheet.append(headers)

        # Apply styles to the header
        for cell in sheet["1"]:
            cell.fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
            cell.font = Font(bold=True)

        workbook.save(str(customer_id) + "Report.xlsx")

        driver.get(secure_messaging_page_getURL(customer_id))
        wait_to_load(driver, 300)
        print("Opened Secure Messaging ")

        # Page loaded assertion
        try:
            dropdown_icon = driver.find_element(By.XPATH, dropdown_icon_xpath)
            contact_log_link = driver.find_element(By.XPATH, contact_log_link_xpath)
            action_click(driver, dropdown_icon)
            driver.execute_script("arguments[0].scrollIntoView();", contact_log_link)
            action_click(driver, contact_log_link)
            start = datetime.now()
            print("Clicked on Contact Log")
            try:
                banner_remove = driver.find_element(By.XPATH, banner_remove_xpath)
                action_click(driver, banner_remove)
            except NoSuchElementException as e:
                print(e)
            # wait_to_load(driver)
            timeout = 300
            try:
                WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, first_column_xpath)))
                # records = len(driver.find_elements(By.XPATH, records_xpath))
                print("Opened Contact Log")
                end = datetime.now()
            except NoSuchElementException:
                print("Page Failed to Load")

            print("Time taken to load ", (end - start).total_seconds(), "s")
            if ((end - start).total_seconds() > 60):
                test_case1_status = ["Page Load", "Verify if Page is loading within proper time ",
                                     "FAIL(Too much time to load)", (end - start).total_seconds()]
                sheet.append(test_case1_status)
            else:
                test_case1_status = ["Page Load", "Verify if Page is loading within proper time ", "PASS",
                                     (end - start).total_seconds()]
                sheet.append(test_case1_status)
        except (
        NoSuchElementException, ElementClickInterceptedException, ElementNotInteractableException, TimeoutException):
            test_case1_status = ["Page Load", "Page failed to load", "FAIL", driver.current_url]
            sheet.append(test_case1_status)

            # Apply conditional formatting
        for cell in sheet["C"][1:]:
            if "FAIL" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
            elif "PASS" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background

        workbook.save(str(customer_id) + "Report.xlsx")

        # Column header assertion - Verifies if column is missing/added
        column_list = driver.find_elements(By.XPATH, columns_xpath)
        actual_columns = []
        for column in column_list:
            actual_columns.append(column.get_attribute("innerHTML"))
        print(actual_columns)
        #
        column_header_match = False
        expected_columns = ['&nbsp;', 'Patient', 'PCP', 'Practice', 'Encounter Date', 'Route', 'Contacted By',
                            'With Whom', 'Encounter Status', 'Encounter Note', 'Created By', 'Measure Status',
                            'Measure Note', 'DOS', 'Compliance Status', 'Health Plan', 'Organization', '&nbsp;']
        if (actual_columns == expected_columns):
            print("Column set matched ! ")
            column_header_match = True
            test_case2_status = ["Column headers", "Verify if columns are not missing or changed ", "PASS",
                                 "No of columns present " + str(len(expected_columns) - 2)]
            sheet.append(test_case2_status)
        else:
            column_header_match = False
            test_case2_status = ["Column headers", "Verify if columns are not missing or changed ", "FAIL",
                                 "No of columns displayed " + str(len(actual_columns) - 2)]
            sheet.append(test_case2_status)
            print("Column set mismatch : Please check expected and actual columns")

        for cell in sheet["C"][1:]:
            if "FAIL" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
            elif "PASS" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background

        workbook.save(str(customer_id) + "Report.xlsx")

        # No of Records Assertion
        records = len(driver.find_elements(By.XPATH, records_xpath))
        print("No of records ", records)
        # check next button enabled or not
        try:
            next = driver.find_element(By.XPATH, next_xpath)
            number_displayed = driver.find_element(By.XPATH, no_of_records_xpath).get_attribute("innerHTML")
            previous = driver.find_element(By.XPATH, previous_xpath)
        except Exception as e:
            print(e)
        # Column Values assertion : # will check if the column header are passed

        # extract value of each column in a separate list

        date_validation_encounter_date = []
        date_validation_dos = []
        if (column_header_match == True):

            # extract values in Encounter Date
            try:
                encounter_dates_values = driver.find_elements(By.XPATH, encounter_dos_xpath)
                for encounter_date in encounter_dates_values:
                    date_validation_encounter_date.append(encounter_date.get_attribute("innerHTML"))

                # extract values in dos
                dos_dates_values = driver.find_elements(By.XPATH, dos_dos_xpath)
                for dos_date in dos_dates_values:
                    date_validation_dos.append(dos_date.get_attribute("innerHTML"))

                print(date_validation_encounter_date)
                print(date_validation_dos)

            except NoSuchElementException as e:
                print("Cant extract column values")

            # verify in each of the format is valid , if all passes good, otherwise pass the record number and failed value

            date_validation_encounter_date_status = "PASS"
            for item in date_validation_encounter_date:
                if (is_valid_datetime(item)):
                    continue
                else:
                    date_validation_encounter_date_status = "FAIL"
                    break

            # add the addition to sheet

            if (date_validation_encounter_date_status == "PASS"):
                test_case8_status = ["Date Format Validation ",
                                     "Verify if Date format is correct for Encounter DOS column ", "PASS",
                                     "Date Format is correctly displayed"]

            else:
                test_case8_status = ["Date Format Validation ",
                                     "Verify if Date format is correct for Encounter DOS column ", "FAIL",
                                     "Date Format is incorrectly displayed"]

            # do the same for DOS column

            dos_date_status = "PASS"
            for item in date_validation_dos:
                if (is_valid_datetime(item) or len(item) == 0):
                    continue
                else:
                    dos_date_status = "FAIL"
                    break

            # add the addition to sheet

            if (dos_date_status == "PASS"):
                test_case9_status = ["Date Format Validation ",
                                     "Verify if Date format is correct for DOS column ", "PASS",
                                     "Date Format is correctly displayed"]

            else:
                test_case9_status = ["Date Format Validation ",
                                     "Verify if Date format is correct for  DOS column ", "FAIL",
                                     "Date Format is incorrectly displayed"]

            sheet.append(test_case8_status)
            sheet.append(test_case9_status)

            for cell in sheet["C"][1:]:
                if "FAIL" in str(cell.value).upper():
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                            fill_type="solid")  # Red background
                elif "PASS" in str(cell.value).upper():
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                            fill_type="solid")  # Green background

            workbook.save(str(customer_id) + "Report.xlsx")

        # Pagination Test Case
        pagination_possible = False
        if (next.is_enabled() and records == 40):
            pagination_possible = True
            print("No of records is more than 40")
            test_case3_status = ["Number of Records ", "No of records is more than 40", "PASS", number_displayed]
            sheet.append(test_case3_status)
        elif records < 40:
            test_case3_status = ["Number of Records ", "No of records is less than 40", "PASS", number_displayed]
            sheet.append(test_case3_status)
            print("No of Records is less than 40")
        else:
            test_case3_status = ["Next is disabled : Please check ", "FAIL",
                                 number_displayed + " " + driver.current_url]
            sheet.append(test_case3_status)
            print("Check Pagination")

        for cell in sheet["C"][1:]:
            if "FAIL" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
            elif "PASS" in str(cell.value).upper():
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background

        workbook.save(str(customer_id) + "Report.xlsx")
        # Check if next is enabled , navigate more than 5 times, compare the record and specify the number of pages checked
        # Check Pagination
        test_case4_status = []
        if (pagination_possible):
            # pages_visited=1
            # next_element=driver.find_element(By.XPATH,next_xpath)
            # footer_text_element=driver.find_element(By.XPATH,no_of_records_xpath)
            # footer_text_original=footer_text_element.get_attribute("innerHTML")
            #
            # #Keep clicking Next and find the number of times it is clicked , when loader appears stop clicking next
            #
            # element_clicked=False
            # no_of_time_clicked=0
            # while not element_clicked:
            #     try:
            #         # Find the element
            #         element = WebDriverWait(driver, 10).until(
            #             EC.presence_of_element_located((By.XPATH, next_xpath))
            #         )
            #
            #         # Click the element
            #         element.click()
            #         loader_found=False
            #         try:
            #             loader_element_class = 'ajax_preloader'
            #             driver.find_element(By.CLASS_NAME,loader_element_class)
            #             loader_found=True
            #         except Exception as e:
            #             pass
            #         if(loader_found==True):
            #             break
            #         # Set the flag to True to break out of the loop
            #         element_clicked = True
            #         no_of_time_clicked=no_of_time_clicked+1
            #
            #     except (
            #             ElementClickInterceptedException, ElementNotInteractableException,
            #             ElementNotInteractableException,
            #             StaleElementReferenceException):
            #         # If the element is stale, continue the loop
            #         continue
            # print("Next is clicked ",no_of_time_clicked)
            #
            #
            #
            # if(next_element.is_enabled()):
            #     print("Next is enabled")
            #     action_click(driver, next_element)
            #     element_clicked = False
            #
            #
            #     wait_to_load(driver, 300)
            # while(pages_visited<7 and next_element.is_enabled()):
            #     wait_to_load(driver, 300)
            #     element_clicked = False
            #
            #     while not element_clicked:
            #         try:
            #             # Find the element
            #             element = WebDriverWait(driver, 10).until(
            #                 EC.presence_of_element_located((By.XPATH, next_xpath))
            #             )
            #
            #             # Click the element
            #             element.click()
            #
            #             # Set the flag to True to break out of the loop
            #             element_clicked = True
            #
            #         except (
            #         ElementClickInterceptedException, ElementNotInteractableException, ElementNotInteractableException,
            #         StaleElementReferenceException):
            #             # If the element is stale, continue the loop
            #             continue
            #     print("Next is clicked ")
            #
            #     pages_visited=pages_visited+1
            #     footer_text_final=footer_text_element.get_attribute("innerHTML")
            # print("Pages visited ", pages_visited)
            # print("Footer Text original", footer_text_original)
            # print("Footer Text last", footer_text_final)

            test_case4_status = ["Pagination Verification ", "Not implemented yet",]

        else:
            test_case4_status = ["Pagination Verification ", "Verify if user can navigate more than 1 page",
                                 "No of records is less than 40 , check for another customer"]
        sheet.append(test_case4_status)

    #if next is disabled within 5 times, compare the record




    for cell in sheet["C"][1:]:
        if "FAIL" in str(cell.value).upper():
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
        elif "PASS" in str(cell.value).upper():
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background

    workbook.save(str(customer_id)+"Report.xlsx")

    #Check Filter
    filter_pass=False

    if(records>1):
        try:
            member_name_all_list=[]
            member_names=driver.find_elements(By.XPATH,name_column_xpath)
            for member_name in member_names:
                member_name_all_list.append(member_name.get_attribute("innerHTML"))
            member_name_set=set(member_name_all_list)
            member_name_list=list(member_name_set)
            print(member_name_list)
            #extract two patient name
            i=1
            while(i<=2):
                #Click on filter
                filter=driver.find_element(By.XPATH,filter_xpath)
                action_click(driver,filter)
                #Apply Patient Name
                patient_name_placeholder=driver.find_element(By.XPATH,patient_name_placeholder_xpath)
                driver.execute_script("arguments[0].scrollIntoView(true);", patient_name_placeholder)
                patient_name_placeholder.clear()
                patient_name_placeholder.send_keys(member_name_list[i-1])
                print("checking for",member_name_list[i-1])
                #Click on Apply

                apply=driver.find_element(By.XPATH,apply_xpath)
                driver.execute_script("arguments[0].scrollIntoView(true);", apply)
                action_click(driver,apply)
                print("Applied filter")
                #count number of records
                start_2 = datetime.now()
                time.sleep(5)
                WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.XPATH, no_of_records_xpath)))
                print("Opened Filtered List")
                end_2 = datetime.now()
                print("Time taken to Filter ", (end_2 - start_2).total_seconds()-5, "s")

                records_2 = len(driver.find_elements(By.XPATH, records_xpath))
                print("No of records after filter ", records_2)
                if(records_2==1):
                    try:
                        no_actual_record = driver.find_element(By.XPATH, no_actual_record_xpath)
                        filter_pass=False
                        break
                    except NoSuchElementException as e:
                        filter_pass = True


                filter_pass=True
                if(records_2==0):
                    filter_pass=False
                    break
                #decide result
                i=i+1
            #enter into
        except Exception as e :
            print(str(e))
    if(filter_pass==True and records>1):
        test_case5_status=["Filter","Verify if Filter for Patient Name is working properly","PASS","Patient name filter working",]
    elif(filter_pass==True and records==1):
        test_case5_status = ["Filter","Verify if Filter for Patient Name is working properly ", "No records ", "Check different customer " ]
    elif (filter_pass==False):
        test_case5_status = ["Filter","Verify if Filter for Patient Name is working properly", "FAIL","Check manually " +driver.current_url]

    sheet.append(test_case5_status)
    for cell in sheet["C"][1:]:
        if "FAIL" in str(cell.value).upper():
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
        elif "PASS" in str(cell.value).upper():
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background

    workbook.save(str(customer_id)+"Report.xlsx")
    print("File saved as "+str(customer_id)+"Report.xlsx")


    #export validation

    #click on download icon

    # try:
    #     #click on export icon
    #     export=driver.find_element(By.XPATH,file_download)
    #     action_click(driver,export)
    #     print("clicked on Export icon")
    #     #check if list is visible
    #     WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,visible_UL)))
    #     #click on Export all to csv
    #     export_to_csv=driver.find_element(By.XPATH,link_to_click)
    #     action_click(driver,export_to_csv)
    #     print("Able to see Export all to csv option")
    #     #check if alert is present
    #     WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, stickets_alert)))
    #     #accept the alert
    #     accept = driver.find_element(By.XPATH,yes_accept)
    #     action_click(driver, accept)
    #     print("Accepted the alert message")
    #     #wait for toast message
    #
    #     # Define the polling interval in seconds
    #     polling_interval = 1
    #
    #     # Define the maximum number of polling attempts
    #     max_attempts = 10
    #
    #     # Poll for the presence of an element
    #     for attempt in range(max_attempts):
    #         try:
    #             element = WebDriverWait(driver, polling_interval).until(
    #                 EC.presence_of_element_located((By.XPATH, toast_message_appears)))
    #             print("Toast message found!")
    #             break  # Exit the loop if element is found
    #         except TimeoutException:
    #             print(f"Attempt {attempt + 1}: Toast message not found yet, retrying...")
    #     else:
    #         print("Toast message not found after maximum attempts.")
    #     #write to sheets
    #     test_case7_status = ["Export", "Verify if user can put the files on download", "PASS","Successful Toast message appears for user" ]
    #     sheet.append(test_case7_status)
    #
    #
    # except (TimeoutException,NoSuchElementException,ElementClickInterceptedException,ElementNotInteractableException) as e:
    #     test_case7_status = ["Export", "Verify if user can put the files on download", "FAIL",
    #                          "User is unable to put the files to download - Check flow manually "]
    #     print("Exception occurred in download "+str(e))
    #
    #Extract the text before notification
    # number_of_notification=int(driver.find_element(By.XPATH,notification_icon).get_attribute("innerHTML"))
    # #refresh the page
    # driver.refresh()
    # wait_to_load(driver)
    # #extract text after refresh
    # number_of_notification_2 = int(driver.find_element(By.XPATH, notification_icon).get_attribute("innerHTML"))
    # #check the difference
    # print("notification Difference"+str(number_of_notification_2-number_of_notification))
    # #Click on download csv option
    # action_click(driver,driver.find_element(By.XPATH,notification_panel))
    # action_click(driver,driver.find_element(By.XPATH,notification_received))
    # # Switch to the newly opened tab
    # driver.switch_to.window(driver.window_handles[1])
    # wait_to_load(driver)
    # # Now you can interact with the elements on the new tab
    # print(driver.title)  # Print the title of the new tab
    #
    # action_click(driver,driver.find_element(By.XPATH,link_to_download_click))
    #
    # # Close the new tab (optional)
    # driver.close()
    #
    # # Switch back to the original tab
    # driver.switch_to.window(driver.window_handles[0])
    # #Download the csv
    #
    # # for cell in sheet["C"][1:]:
    # #     if "FAIL" in str(cell.value).upper():
    # #         cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red background
    # #     elif "PASS" in str(cell.value).upper():
    # #         cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green background
    #
    # workbook.save(str(customer_id)+"Report.xlsx")
    # print("File saved as "+str(customer_id)+"Report.xlsx")


    print("Verification continued ")



def action_click(driver,element):
    try:
        element.click()
    except (ElementNotInteractableException, ElementClickInterceptedException,StaleElementReferenceException):
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        driver.execute_script("arguments[0].click();", element)


def check_chartlist_export(driver,customer_id):
    hamburger_icon="//i[text()='menu']"
    supplemental_data_link_xpath="//li[@class='chart_chase_list_type' and @data-list-type='1']//a"
    column_header_xpath="//th[2]"
    filter_list_xpath = "//i[text()=\"filter_list\"]"
    new_creation_date_filter_from_xpath = "//input[@name='chart_chase_uploaded_from']"
    new_creation_date_filter_to_xpath = "//input[@name='chart_chase_uploaded_to']"
    apply_xpath = "//a[text()=\"Apply\"]"
    footer_xpath="//div[@class='dataTables_info']"
    export_icon_xpath="//a[@data-tooltip=\"Export\"]"
    export_list_xpath="//a[text()='Export all to CSV ']"
    # export_option_xpath=

    #make a directory of customer_id in  C://ChartListExports//customer_id
    # Path to the directory
    directory_path = f"C:/ChartListExports/{customer_id}"

    # Create the directory
    os.makedirs(directory_path, exist_ok=True)

    # Check if the directory was created
    if os.path.exists(directory_path):
        print(f"Directory '{directory_path}' created successfully.")
    else:
        print(f"Failed to create directory '{directory_path}'.")

    #Open Registry page for customer
    customer_list_url = []
    sm_customer_id = customer_id
    session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
        sm_customer_id) + '&orgId=' + str(sm_customer_id) + '&vgpId=' + str(sm_customer_id) + '&vpId=' + str(
        sm_customer_id)
    encoded_string = base64.b64encode(session_var.encode('utf-8'))
    customer_list_url.append(encoded_string)
    for idx, val in enumerate(customer_list_url):
        url = ("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))

    driver.get(url)
    #open supplemental data chart list
    wait_to_load(driver, 300)
    action_click(driver,driver.find_element(By.XPATH,hamburger_icon))
    action_click(driver,driver.find_element(By.XPATH,supplemental_data_link_xpath))


    #wait for page to load

    wait_to_load(driver,300)
    timeout_for_column_headers=20
    WebDriverWait(driver, timeout_for_column_headers).until(EC.visibility_of_element_located((By.XPATH, column_header_xpath)))

    time_delta=5

    while(1):
        # Set date filter
        # Get the current date
        current_date = datetime.now()
        formatted_date_to = current_date.strftime("%m/%d/%Y")
        date_from = current_date - timedelta(days=time_delta)
        formatted_date_from = date_from.strftime("%m/%d/%Y")

        WebDriverWait(driver, timeout_for_column_headers).until(
            EC.element_to_be_clickable((By.XPATH, filter_list_xpath)))
        # first apply date filter
        action_click(driver, driver.find_element(By.XPATH, filter_list_xpath))
        time.sleep(2)
        driver.find_element(By.XPATH, new_creation_date_filter_from_xpath).clear()
        driver.find_element(By.XPATH, new_creation_date_filter_from_xpath).send_keys(formatted_date_from)

        driver.find_element(By.XPATH, new_creation_date_filter_to_xpath).clear()
        driver.find_element(By.XPATH, new_creation_date_filter_to_xpath).send_keys(formatted_date_to)

        action_click(driver, driver.find_element(By.XPATH, apply_xpath))

        wait_to_load(driver, 300)
        time.sleep(2)
        WebDriverWait(driver, timeout_for_column_headers).until(
            EC.visibility_of_element_located((By.XPATH, footer_xpath)))

        # check number of entries
        footer_text = driver.find_element(By.XPATH, footer_xpath).get_attribute("innerHTML")
        # Find the index of "of" and "entries"
        index_of_of = footer_text.find("of")
        index_of_entries = footer_text.find("entries")

        # Extract the number between "of" and "entries"
        number_between_of_and_entries = int(footer_text[index_of_of + 3:index_of_entries].strip().replace(",", ""))

        print("No of entries ", number_between_of_and_entries)
        if (number_between_of_and_entries > 300 and number_between_of_and_entries < 2500):

            print(f"Optimal entries present for {time_delta} days before current date ")
            break
        if (number_between_of_and_entries < 300):
            time_delta = time_delta +5
        if (number_between_of_and_entries > 2500):
            time_delta = time_delta - 2

    #download file to C://Downloads//ChartList//CustomerID//__.csv
    download_successful=False
    try:
        action_click(driver, driver.find_element(By.XPATH, export_icon_xpath))
        action_click(driver, driver.find_element(By.XPATH, export_list_xpath))
        download_successful=True
        time.sleep(5)


    except Exception as e:
        print("Error occurred in downloading file "+str(e))

    if(download_successful):
        # if above is successful , copy the file
        destination_dir = f"C:/ChartListExports/{customer_id}"

        # Get a list of all files in the download folder
        files = os.listdir(download_dir)

        # Filter out directories from the list
        files = [f for f in files if os.path.isfile(os.path.join(download_dir, f))]

        # Get the latest downloaded file based on modification time
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(download_dir, x)))

        # Path to the latest downloaded file
        latest_file_path = os.path.join(download_dir, latest_file)

        # Path to the destination file
        destination_file_path = os.path.join(destination_dir, latest_file)

        # Create the destination directory if it doesn't exist
        os.makedirs(destination_dir, exist_ok=True)

        # Copy the file to the destination directory
        shutil.move(latest_file_path, destination_file_path)

        print(f"Latest downloaded file '{latest_file}' moved to '{destination_dir}' successfully.")
        return [customer_id,destination_dir,0,0,report]


def execute(environment_name,page_name,customer_selection):
    if(environment_name=="CERT"):
        logout_url="https://cert.cozeva.com/user/logout"
        login_url="https://cert.cozeva.com/user/login"
    if (environment_name == "STAGE"):
        logout_url = "https://stage.cozeva.com/user/logout"
        login_url = "https://stage.cozeva.com/user/login"
    if (environment_name == "PROD"):
        logout_url = "https://www.cozeva.com/user/logout"
        login_url = "https://www.cozeva.com/user/login"
    driver=setup()
    driver.get(logout_url)
    driver.get(login_url)
    creds = get_credentials(environment_name)

    #login
    uname = driver.find_element(By.ID,"edit-name")
    pwd = driver.find_element(By.ID,"edit-pass")
    uname.send_keys(creds[0])
    pwd.send_keys(creds[1])
    driver.find_element(By.ID,"edit-submit").click()
    # reason for login
    WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, "//textarea[@id=\"reason_textbox\"]")))
    actions = ActionChains(driver)
    reason = driver.find_element(By.XPATH,"//textarea[@id=\"reason_textbox\"]")
    actions.click(reason)
    actions.send_keys_to_element(reason, "https://redmine2.cozeva.com/issues/7662 ")
    actions.perform()
    driver.find_element(By.ID,"edit-submit").click()
    print("Logged in")
    # ["1300", "3900" , "1950" , "2200" , "1000"]
    customer_ids=["2200"]
    #open secure messaging
    wb=Workbook()

    if(page_name=="Quality Chart List"):
        check_chartlist_export(driver,customer_ids)


    sample_logger=1
    run_from="CS"
    #throws TimeoutException
    check_contact_log(driver,  wb, sample_logger, run_from ,customer_ids)








    #record page load time




#login to Cozeva

#From a list of customers , create secure messaging link - use customer list and page

#create new report - SecureMessaging-displaying - Link/Data Record/Time to load/Comments -- Time to load will display Time if pass and comments if failed


#record the time for sticket/load  - contact page load - Check Error occurred in each link

#handle exception

#for sticket and contact log - record time to load page

#Check the columns displayed



